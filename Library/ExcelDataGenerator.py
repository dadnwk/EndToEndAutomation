import openpyxl


def data_generator():
    wk = openpyxl.load_workbook("./DataTest/SearchDataTest.xlsx")  # "./DataTest/SearchDataTest.xlsx"
    sh = wk['SearchData']
    r = sh.max_row
    li = []

    for i in range(1, r + 1):
        li1 = []
        un = sh.cell(i, 1)
        up = sh.cell(i, 2)
        li1.insert(0, un.value)
        li1.insert(1, up.value)
        li.insert(i - 1, li1)

    print(li)
    return li


# check if the function works or not
data_generator()
